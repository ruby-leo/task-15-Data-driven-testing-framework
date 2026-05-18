from openpyxl import load_workbook

def get_username_password_from_excel():
    wb = load_workbook('test_data_and_report.xlsx')
    sheet = wb['Sheet1']
    data = list(sheet.values)
    credentials_list = []
    for index, item in enumerate(data[1:], start=2):
        row_data = [index, item[1], item[2]]
        credentials_list.append(row_data)
    return credentials_list

def write_results_back_into_excel(row_num, date, time_of_test, name_of_tester, test_result):
    print("row num is: ", row_num, date, time_of_test, name_of_tester, test_result)
    wb = load_workbook('test_data_and_report.xlsx')
    sheet = wb['Sheet1']
    sheet.cell(row=row_num, column=4).value = date
    sheet.cell(row=row_num, column=5).value = time_of_test
    sheet.cell(row=row_num, column=6).value = name_of_tester
    sheet.cell(row=row_num, column=7).value = test_result
    wb.save('test_data_and_report.xlsx')
    wb.close()